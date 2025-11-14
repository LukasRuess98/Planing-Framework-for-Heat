"""Utilities for exporting scenario inputs and optimization results."""

from __future__ import annotations

from collections import OrderedDict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from numbers import Number
from typing import Iterable, Mapping, MutableMapping, Sequence
import csv
import json
import math
import os
from zipfile import ZIP_DEFLATED, ZipFile

from xml.sax.saxutils import escape

_XML_ESCAPE = {"'": "&apos;"}

from energis.utils.timeseries import TimeSeriesTable

try:  # pragma: no cover - optional dependency
    from openpyxl import Workbook
    HAVE_OPENPYXL = True
except Exception:  # pragma: no cover
    Workbook = None
    HAVE_OPENPYXL = False

__all__ = [
    "write_timeseries_csv",
    "write_excel_workbook",
    "write_scenario_workbook",
    "export_scenario_bundle",
    "HAVE_OPENPYXL",
]


def _ensure_lengths(data: Mapping[str, Sequence[float]], expected: int) -> None:
    for key, values in data.items():
        if len(values) != expected:
            raise ValueError(f"Serie {key!r} hat Länge {len(values)} statt {expected}")


def write_timeseries_csv(
    path: str,
    table: TimeSeriesTable,
    extra: Mapping[str, Sequence[float]],
    *,
    decimal_separator: str = ",",
    alternate_path: str | None = None,
    alternate_decimal_separator: str = ".",
) -> None:
    """Write input and result time series to a semicolon separated CSV file.

    Parameters
    ----------
    path:
        Destination of the main CSV export.
    table:
        The base time series table to be exported.
    extra:
        Additional series that should be exported next to the main table.
    decimal_separator:
        Decimal separator used for *path*. Defaults to a comma which is
        convenient for German Excel installations.
    alternate_path:
        Optional second file that mirrors the exported content but allows
        specifying a different decimal separator. This can be helpful when the
        data should be double-checked in tools that prefer the conventional dot
        separator.
    alternate_decimal_separator:
        Decimal separator used for *alternate_path*. Ignored when no alternate
        path is supplied.
    """

    n = len(table)
    _ensure_lengths(extra, n)

    columns = ["timestamp"] + table.columns + list(extra.keys())
    _write_csv(
        path,
        columns,
        table,
        extra,
        decimal_separator=decimal_separator,
    )

    if alternate_path:
        _write_csv(
            alternate_path,
            columns,
            table,
            extra,
            decimal_separator=alternate_decimal_separator,
        )


def _fmt_value(value: object, *, decimal_separator: str = ",") -> str:
    if value is None:
        return ""

    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return ""
        try:
            decimal_value = _to_decimal(stripped)
        except InvalidOperation:
            return value
        text = _decimal_to_text(decimal_value)
        return _apply_decimal_separator(text, decimal_separator)

    if isinstance(value, Number) and not isinstance(value, bool):
        decimal_value = _to_decimal(value)
        text = _decimal_to_text(decimal_value)
        return _apply_decimal_separator(text, decimal_separator)

    return str(value)


def _write_csv(
    path: str,
    columns: Sequence[str],
    table: TimeSeriesTable,
    extra: Mapping[str, Sequence[float]],
    *,
    decimal_separator: str,
) -> None:
    with open(path, "w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(
            handle,
            delimiter=";",
            quoting=csv.QUOTE_MINIMAL,
            lineterminator="\n",
        )
        writer.writerow(columns)
        for idx, ts in enumerate(table.index):
            base = [ts.isoformat() if isinstance(ts, datetime) else str(ts)]
            base.extend(
                _fmt_value(table[col][idx], decimal_separator=decimal_separator)
                for col in table.columns
            )
            base.extend(
                _fmt_value(extra[name][idx], decimal_separator=decimal_separator)
                for name in extra
            )
            writer.writerow(base)


def _to_decimal(value: Number | Decimal | str) -> Decimal:
    if isinstance(value, Decimal):
        return value

    if isinstance(value, str):
        normalized = value.replace(" ", "").replace(",", ".")
        return Decimal(normalized)

    try:
        return Decimal(str(value))
    except InvalidOperation:
        return Decimal(value)


def _decimal_to_text(value: Decimal) -> str:
    text = format(value.normalize(), "f")
    if "." in text:
        text = text.rstrip("0").rstrip(".")
    return text or "0"


def _apply_decimal_separator(text: str, decimal_separator: str) -> str:
    if decimal_separator not in {",", "."}:
        raise ValueError(
            "decimal_separator must be ',' or '.' to avoid ambiguous CSV numbers"
        )
    if decimal_separator == ".":
        return text
    return text.replace(".", ",")


def _normalize_excel_value(value: object) -> object:
    """Convert arbitrary values into a representation accepted by openpyxl."""

    if value is None:
        return None

    if isinstance(value, (str, Number, bool)):
        return value

    if isinstance(value, datetime):
        return value

    if isinstance(value, date):
        # openpyxl expects datetime for date-like objects
        return datetime.combine(value, datetime.min.time())

    iso_formatter = getattr(value, "isoformat", None)
    if callable(iso_formatter):
        try:
            return iso_formatter()
        except Exception:
            pass

    return str(value)


def _append_table(ws, headers: Sequence[str], rows: Iterable[Sequence[object]]) -> None:
    ws.append([_normalize_excel_value(header) for header in headers])
    for row in rows:
        ws.append([_normalize_excel_value(value) for value in row])


def _excel_safe_value(value: object) -> object:
    if value is None:
        return ""

    if isinstance(value, datetime):
        return value.isoformat(sep=" ")

    if isinstance(value, date):
        return value.isoformat()

    if isinstance(value, Decimal):
        value = float(value)

    if isinstance(value, Number) and not isinstance(value, bool):
        if isinstance(value, float) and not math.isfinite(value):
            return ""
        return value

    if isinstance(value, bool):
        return value

    if isinstance(value, (list, tuple, dict)):
        try:
            return json.dumps(value, ensure_ascii=False)
        except TypeError:
            return str(value)

    return str(value)


def _excel_safe_df(columns: Sequence[str], rows: Iterable[Mapping[str, object]]) -> list[list[object]]:
    safe_rows: list[list[object]] = []
    for row in rows:
        safe_row: list[object] = []
        for col in columns:
            safe_row.append(_excel_safe_value(row.get(col, "")))
        safe_rows.append(safe_row)
    return safe_rows


def _column_letter(idx: int) -> str:
    if idx <= 0:
        return "A"
    result = ""
    while idx:
        idx, remainder = divmod(idx - 1, 26)
        result = chr(65 + remainder) + result
    return result or "A"


def _sheet_xml(rows: Sequence[Sequence[object]]) -> str:
    parts = [
        "<?xml version=\"1.0\" encoding=\"UTF-8\"?>",
        "<worksheet xmlns=\"http://schemas.openxmlformats.org/spreadsheetml/2006/main\">",
        "<sheetData>",
    ]
    for row_idx, row in enumerate(rows, start=1):
        cells: list[str] = []
        for col_idx, value in enumerate(row, start=1):
            if value is None or value == "":
                continue
            text = str(value)
            text = text.replace("\r\n", "\n").replace("\r", "\n")
            text = escape(text, _XML_ESCAPE)
            text = text.replace("\n", "&#10;")
            cell_ref = f"{_column_letter(col_idx)}{row_idx}"
            cells.append(
                f"<c r=\"{cell_ref}\" t=\"inlineStr\"><is><t>{text}</t></is></c>"
            )
        if cells:
            parts.append(f"<row r=\"{row_idx}\">{''.join(cells)}</row>")
        else:
            parts.append(f"<row r=\"{row_idx}\"/>")
    parts.append("</sheetData></worksheet>")
    return "".join(parts)


def _workbook_xml(sheet_names: Sequence[str]) -> str:
    parts = [
        "<?xml version=\"1.0\" encoding=\"UTF-8\"?>",
        "<workbook xmlns=\"http://schemas.openxmlformats.org/spreadsheetml/2006/main\" xmlns:r=\"http://schemas.openxmlformats.org/officeDocument/2006/relationships\">",
        "<sheets>",
    ]
    for idx, name in enumerate(sheet_names, start=1):
        safe_name = escape(str(name), _XML_ESCAPE)
        parts.append(
            f"<sheet name=\"{safe_name}\" sheetId=\"{idx}\" r:id=\"rId{idx}\"/>"
        )
    parts.append("</sheets></workbook>")
    return "".join(parts)


def _workbook_rels_xml(num_sheets: int) -> str:
    parts = [
        "<?xml version=\"1.0\" encoding=\"UTF-8\"?>",
        "<Relationships xmlns=\"http://schemas.openxmlformats.org/package/2006/relationships\">",
    ]
    for idx in range(1, num_sheets + 1):
        parts.append(
            f"<Relationship Id=\"rId{idx}\" Type=\"http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet\" Target=\"worksheets/sheet{idx}.xml\"/>"
        )
    parts.append("</Relationships>")
    return "".join(parts)


def _root_rels_xml() -> str:
    return (
        "<?xml version=\"1.0\" encoding=\"UTF-8\"?>"
        "<Relationships xmlns=\"http://schemas.openxmlformats.org/package/2006/relationships\">"
        "<Relationship Id=\"rId1\" Type=\"http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument\" Target=\"xl/workbook.xml\"/>"
        "<Relationship Id=\"rId2\" Type=\"http://schemas.openxmlformats.org/package/2006/relationships/metadata/core-properties\" Target=\"docProps/core.xml\"/>"
        "<Relationship Id=\"rId3\" Type=\"http://schemas.openxmlformats.org/officeDocument/2006/relationships/extended-properties\" Target=\"docProps/app.xml\"/>"
        "</Relationships>"
    )


def _content_types_xml(num_sheets: int) -> str:
    parts = [
        "<?xml version=\"1.0\" encoding=\"UTF-8\"?>",
        "<Types xmlns=\"http://schemas.openxmlformats.org/package/2006/content-types\">",
        "<Default Extension=\"rels\" ContentType=\"application/vnd.openxmlformats-package.relationships+xml\"/>",
        "<Default Extension=\"xml\" ContentType=\"application/xml\"/>",
        "<Override PartName=\"/xl/workbook.xml\" ContentType=\"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml\"/>",
        "<Override PartName=\"/xl/styles.xml\" ContentType=\"application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml\"/>",
        "<Override PartName=\"/docProps/core.xml\" ContentType=\"application/vnd.openxmlformats-package.core-properties+xml\"/>",
        "<Override PartName=\"/docProps/app.xml\" ContentType=\"application/vnd.openxmlformats-officedocument.extended-properties+xml\"/>",
    ]
    for idx in range(1, num_sheets + 1):
        parts.append(
            f"<Override PartName=\"/xl/worksheets/sheet{idx}.xml\" ContentType=\"application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml\"/>"
        )
    parts.append("</Types>")
    return "".join(parts)


def _styles_xml() -> str:
    return (
        "<?xml version=\"1.0\" encoding=\"UTF-8\"?>"
        "<styleSheet xmlns=\"http://schemas.openxmlformats.org/spreadsheetml/2006/main\">"
        "<fonts count=\"1\"><font><sz val=\"11\"/><name val=\"Calibri\"/></font></fonts>"
        "<fills count=\"1\"><fill><patternFill patternType=\"none\"/></fill></fills>"
        "<borders count=\"1\"><border/></borders>"
        "<cellStyleXfs count=\"1\"><xf/></cellStyleXfs>"
        "<cellXfs count=\"1\"><xf xfId=\"0\"/></cellXfs>"
        "<cellStyles count=\"1\"><cellStyle name=\"Normal\" xfId=\"0\" builtinId=\"0\"/></cellStyles>"
        "</styleSheet>"
    )


def _app_xml(sheet_names: Sequence[str]) -> str:
    count = len(sheet_names)
    titles = "".join(
        f"<vt:lpstr>{escape(str(name), _XML_ESCAPE)}</vt:lpstr>" for name in sheet_names
    )
    return (
        "<?xml version=\"1.0\" encoding=\"UTF-8\"?>"
        "<Properties xmlns=\"http://schemas.openxmlformats.org/officeDocument/2006/extended-properties\" xmlns:vt=\"http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes\">"
        "<Application>energis</Application>"
        "<HeadingPairs><vt:vector size=\"2\" baseType=\"variant\">"
        "<vt:variant><vt:lpstr>Worksheets</vt:lpstr></vt:variant>"
        f"<vt:variant><vt:i4>{count}</vt:i4></vt:variant>"
        "</vt:vector></HeadingPairs>"
        "<TitlesOfParts><vt:vector baseType=\"lpstr\" size=\"{count}\">"
        f"{titles}"
        "</vt:vector></TitlesOfParts>"
        "</Properties>"
    )


def _core_xml() -> str:
    now = datetime.utcnow().replace(microsecond=0).isoformat() + "Z"
    return (
        "<?xml version=\"1.0\" encoding=\"UTF-8\"?>"
        "<cp:coreProperties xmlns:cp=\"http://schemas.openxmlformats.org/package/2006/metadata/core-properties\" xmlns:dc=\"http://purl.org/dc/elements/1.1/\" xmlns:dcterms=\"http://purl.org/dc/terms/\" xmlns:xsi=\"http://www.w3.org/2001/XMLSchema-instance\">"
        "<dc:creator>energis</dc:creator>"
        "<cp:lastModifiedBy>energis</cp:lastModifiedBy>"
        f"<dcterms:created xsi:type=\"dcterms:W3CDTF\">{now}</dcterms:created>"
        f"<dcterms:modified xsi:type=\"dcterms:W3CDTF\">{now}</dcterms:modified>"
        "<cp:revision>1</cp:revision>"
        "</cp:coreProperties>"
    )


def _write_simple_xlsx(path: str, sheets: Mapping[str, Sequence[Sequence[object]]]) -> None:
    directory = os.path.dirname(path)
    if directory:
        os.makedirs(directory, exist_ok=True)

    sheet_names = list(sheets.keys())
    with ZipFile(path, "w", ZIP_DEFLATED) as handle:
        handle.writestr("[Content_Types].xml", _content_types_xml(len(sheet_names)))
        handle.writestr("_rels/.rels", _root_rels_xml())
        handle.writestr("docProps/app.xml", _app_xml(sheet_names))
        handle.writestr("docProps/core.xml", _core_xml())
        handle.writestr("xl/workbook.xml", _workbook_xml(sheet_names))
        handle.writestr("xl/_rels/workbook.xml.rels", _workbook_rels_xml(len(sheet_names)))
        handle.writestr("xl/styles.xml", _styles_xml())
        for idx, (name, rows) in enumerate(sheets.items(), start=1):
            handle.writestr(f"xl/worksheets/sheet{idx}.xml", _sheet_xml(rows))


def _build_meta_sheet(meta_sections: Mapping[str, Mapping[str, object]] | None) -> list[list[object]]:
    if not meta_sections:
        return [["no metadata available"]]

    rows: list[list[object]] = []
    for section, entries in meta_sections.items():
        rows.append([_excel_safe_value(section)])
        if isinstance(entries, Mapping):
            for key, value in entries.items():
                rows.append([
                    _excel_safe_value(key),
                    _excel_safe_value(value),
                ])
        else:
            rows.append(["value", _excel_safe_value(entries)])
        rows.append([])
    return rows


def _build_timeseries_sheet(
    sections: Sequence[Mapping[str, object]] | None,
) -> list[list[object]]:
    if not sections:
        return [["no timeseries available"]]

    rows: list[list[object]] = []
    for section in sections:
        label = _excel_safe_value(section.get("label", "timeseries"))
        timestamps = list(section.get("timestamps", []))
        series = section.get("series")
        rows.append([label])
        if not timestamps or not isinstance(series, Mapping) or not series:
            rows.append(["no data"])
            rows.append([])
            continue

        value_lists: "OrderedDict[str, list[object]]" = OrderedDict()
        for name, values in series.items():
            key = str(name)
            seq = list(values)
            if len(seq) < len(timestamps):
                seq.extend([""] * (len(timestamps) - len(seq)))
            elif len(seq) > len(timestamps):
                seq = seq[: len(timestamps)]
            value_lists[key] = seq

        headers = ["timestamp"] + list(value_lists.keys())
        rows.append([_excel_safe_value(header) for header in headers])

        data_rows = []
        for idx, ts in enumerate(timestamps):
            row: "OrderedDict[str, object]" = OrderedDict()
            row["timestamp"] = ts
            for name, seq in value_lists.items():
                row[name] = seq[idx]
            data_rows.append(row)

        rows.extend(_excel_safe_df(headers, data_rows))
        rows.append([])

    return rows or [["no timeseries available"]]


def _build_costs_sheet(costs: Mapping[str, Mapping[str, object]] | None) -> list[list[object]]:
    if not costs:
        return [["no costs available"]]

    rows: list[list[object]] = []
    for section, entries in costs.items():
        rows.append([_excel_safe_value(section)])
        if isinstance(entries, Mapping):
            for key, value in entries.items():
                rows.append([
                    _excel_safe_value(key),
                    _excel_safe_value(value),
                ])
        else:
            rows.append(["value", _excel_safe_value(entries)])
        rows.append([])
    return rows


def _build_design_sheet(design: Mapping[str, object] | None) -> list[list[object]]:
    if not design:
        return [["no design data available"]]

    rows: list[list[object]] = []
    heat_pumps = design.get("heat_pumps") if isinstance(design, Mapping) else None
    if isinstance(heat_pumps, Mapping) and heat_pumps:
        rows.append([_excel_safe_value("heat_pumps")])
        rows.append([
            _excel_safe_value("id"),
            _excel_safe_value("capacity_mw"),
            _excel_safe_value("build_binary"),
        ])
        for hp_id, metrics in heat_pumps.items():
            if isinstance(metrics, Mapping):
                rows.append(
                    [
                        _excel_safe_value(hp_id),
                        _excel_safe_value(metrics.get("capacity_mw")),
                        _excel_safe_value(metrics.get("build_binary")),
                    ]
                )
        rows.append([])

    storage = design.get("storage") if isinstance(design, Mapping) else None
    if isinstance(storage, Mapping) and storage:
        rows.append([_excel_safe_value("storage")])
        rows.append([
            _excel_safe_value("name"),
            _excel_safe_value("capacity_mwh"),
            _excel_safe_value("power_mw"),
            _excel_safe_value("build_binary"),
        ])
        rows.append(
            [
                _excel_safe_value(storage.get("name")),
                _excel_safe_value(storage.get("capacity_mwh")),
                _excel_safe_value(storage.get("power_mw")),
                _excel_safe_value(storage.get("build_binary")),
            ]
        )
        rows.append([])

    if not rows:
        rows.append(["no design data available"])
    return rows


def write_scenario_workbook(
    path: str,
    *,
    meta_sections: Mapping[str, Mapping[str, object]] | None = None,
    timeseries_sections: Sequence[Mapping[str, object]] | None = None,
    cost_sections: Mapping[str, Mapping[str, object]] | None = None,
    design: Mapping[str, object] | None = None,
) -> None:
    sheets: "OrderedDict[str, list[list[object]]]" = OrderedDict()
    sheets["Meta"] = _build_meta_sheet(meta_sections)
    sheets["Timeseries"] = _build_timeseries_sheet(timeseries_sections)
    sheets["Costs"] = _build_costs_sheet(cost_sections)
    sheets["Design"] = _build_design_sheet(design)
    _write_simple_xlsx(path, sheets)


def export_scenario_bundle(
    outdir: str,
    *,
    meta_sections: Mapping[str, Mapping[str, object]] | None = None,
    timeseries_sections: Sequence[Mapping[str, object]] | None = None,
    cost_sections: Mapping[str, Mapping[str, object]] | None = None,
    design: Mapping[str, object] | None = None,
    manifest: Mapping[str, object] | None = None,
) -> Mapping[str, object | None]:
    os.makedirs(outdir, exist_ok=True)

    workbook_path = os.path.join(outdir, "scenario.xlsx")
    write_scenario_workbook(
        workbook_path,
        meta_sections=meta_sections,
        timeseries_sections=timeseries_sections,
        cost_sections=cost_sections,
        design=design,
    )

    manifest_data: "OrderedDict[str, object]" = OrderedDict()
    if manifest:
        for key, value in manifest.items():
            manifest_data[key] = value

    files_entry: MutableMapping[str, object]
    raw_files = manifest_data.get("files")
    if isinstance(raw_files, MutableMapping):
        files_entry = raw_files
    else:
        files_entry = OrderedDict()
        manifest_data["files"] = files_entry

    files_entry["scenario_xlsx"] = os.path.basename(workbook_path)

    design_path: str | None = None
    if isinstance(design, Mapping) and (
        (isinstance(design.get("heat_pumps"), Mapping) and design.get("heat_pumps"))
        or (isinstance(design.get("storage"), Mapping) and design.get("storage"))
    ):
        design_path = os.path.join(outdir, "pf_design.json")
        with open(design_path, "w", encoding="utf-8") as handle:
            json.dump(design, handle, indent=2)
        files_entry["pf_design_json"] = os.path.basename(design_path)

    manifest_path = os.path.join(outdir, "manifest.json")
    with open(manifest_path, "w", encoding="utf-8") as handle:
        json.dump(manifest_data, handle, indent=2)

    return {
        "scenario_xlsx": workbook_path,
        "pf_design_json": design_path,
        "manifest_json": manifest_path,
    }


def write_excel_workbook(
    path: str,
    table: TimeSeriesTable,
    series: Mapping[str, Sequence[float]],
    summary_sections: Mapping[str, Mapping[str, object]],
    metadata_sections: Mapping[str, Mapping[str, object]] | None = None,
) -> None:
    """Create an Excel workbook containing inputs, results and summary tables."""

    if not HAVE_OPENPYXL:
        raise RuntimeError(
            "openpyxl ist nicht installiert – Excel-Export ist daher nicht möglich. Bitte openpyxl nachinstallieren."
        )

    n = len(table)
    _ensure_lengths(series, n)

    wb = Workbook()

    input_ws = wb.active
    input_ws.title = "input_timeseries"
    _append_table(
        input_ws,
        ["timestamp"] + table.columns,
        (
            [table.index[i].isoformat() if isinstance(table.index[i], datetime) else str(table.index[i])] +
            [table[col][i] for col in table.columns]
            for i in range(n)
        ),
    )
    input_ws.freeze_panes = "B2"

    results_ws = wb.create_sheet("results_timeseries")
    series_headers = ["timestamp"] + list(series.keys())
    _append_table(
        results_ws,
        series_headers,
        (
            [table.index[i].isoformat() if isinstance(table.index[i], datetime) else str(table.index[i])] +
            [series[name][i] for name in series]
            for i in range(n)
        ),
    )
    results_ws.freeze_panes = "B2"

    summary_ws = wb.create_sheet("summary")
    for section, metrics in summary_sections.items():
        summary_ws.append([_normalize_excel_value(section)])
        for key, value in metrics.items():
            summary_ws.append(
                [_normalize_excel_value(key), _normalize_excel_value(value)]
            )
        summary_ws.append([])

    if metadata_sections:
        meta_ws = wb.create_sheet("metadata")
        for section, entries in metadata_sections.items():
            meta_ws.append([_normalize_excel_value(section)])
            for key, value in entries.items():
                if isinstance(value, (dict, list)):
                    normalized_value = json.dumps(value, ensure_ascii=False)
                else:
                    normalized_value = _normalize_excel_value(value)
                meta_ws.append(
                    [_normalize_excel_value(key), _normalize_excel_value(normalized_value)]
                )
            meta_ws.append([])

    wb.save(path)
